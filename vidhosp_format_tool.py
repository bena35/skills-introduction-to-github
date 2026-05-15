#!/usr/bin/env python3
"""Validation/correction VID-HOSP à largeur fixe selon un classeur de spécifications."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Le module 'openpyxl' est requis. Installez-le avec: pip install openpyxl"
    ) from exc


@dataclass(frozen=True)
class FieldSpec:
    name: str
    width: int
    start: int
    end: int
    required: str


def normalize(text: str) -> str:
    """Normalize text for stable header matching across accents and punctuation."""
    text = (text or "").strip().lower()
    replacements = {
        "é": "e",
        "è": "e",
        "ê": "e",
        "ë": "e",
        "à": "a",
        "â": "a",
        "î": "i",
        "ï": "i",
        "ô": "o",
        "ö": "o",
        "ù": "u",
        "û": "u",
        "ü": "u",
        "ç": "c",
        "°": "",
        "n°": "numero",
        "№": "numero",
        "’": "'",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def format_cell_value(value) -> str:
    """Convert an Excel cell value to a one-line string representation."""
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%d%m%Y")
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def fit_width(value: str, width: int) -> str:
    """Return value truncated or right-padded with spaces to exactly `width` chars."""
    value = value or ""
    if len(value) > width:
        return value[:width]
    return value.ljust(width)


def load_vid_hosp_spec(format_xlsx: Path, sheet_name: str) -> List[FieldSpec]:
    """Load fixed-width field specifications from the VID-HOSP format sheet."""
    wb = load_workbook(format_xlsx, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet introuvable dans le fichier format: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_index = None
    for i, row in enumerate(rows):
        row_vals = [str(c).strip() if c is not None else "" for c in row[:4]]
        if row_vals[:4] == ["Nom", "Taille", "Début", "Fin"]:
            header_index = i
            break

    if header_index is None:
        raise ValueError("En-tête du format introuvable (Nom/Taille/Début/Fin).")

    specs: List[FieldSpec] = []
    for row in rows[header_index + 1 :]:
        name, size, start, end = row[0], row[1], row[2], row[3]
        required = row[6] if len(row) > 6 else ""

        if name is None:
            continue

        if all(v is not None for v in (size, start, end)):
            try:
                specs.append(
                    FieldSpec(
                        name=str(name).strip(),
                        width=int(size),
                        start=int(start),
                        end=int(end),
                        required=str(required or "").strip(),
                    )
                )
            except (TypeError, ValueError):
                continue

    if not specs:
        raise ValueError("Aucun champ exploitable trouvé dans l'onglet VID-HOSP.")

    return sorted(specs, key=lambda x: x.start)


def total_width(specs: List[FieldSpec]) -> int:
    """Return expected fixed-width record length from field definitions."""
    if not specs:
        raise ValueError("Aucune spécification de champ fournie.")
    return max(f.end for f in specs)


def build_line_from_fixed(raw_line: str, specs: List[FieldSpec]) -> str:
    """Rebuild one fixed-width line according to field start/end definitions."""
    out = []
    for f in specs:
        start_idx = f.start - 1
        end_idx = f.end
        segment = raw_line[start_idx:end_idx] if start_idx < len(raw_line) else ""
        out.append(fit_width(segment, f.width))
    return "".join(out)


def fix_txt_lines(specs: List[FieldSpec], src_txt: Path, dst_txt: Path, report_csv: Path | None) -> None:
    """Fix each input TXT line to the expected fixed-width layout and write a report."""
    expected_len = total_width(specs)
    reports = []

    with src_txt.open("r", encoding="utf-8", errors="replace") as fin, dst_txt.open(
        "w", encoding="utf-8", newline="\n"
    ) as fout:
        for i, line in enumerate(fin, start=1):
            raw = line.rstrip("\r\n")
            fixed = build_line_from_fixed(raw, specs)
            changed = raw != fixed
            reports.append((i, len(raw), len(fixed), "YES" if changed else "NO"))
            fout.write(fixed + "\n")

    if report_csv:
        with report_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["line", "input_length", "output_length", "changed"])
            writer.writerows(reports)

    print(f"OK: {src_txt} -> {dst_txt} (longueur attendue={expected_len})")


def build_header_map(header_cells: Iterable) -> dict[str, int]:
    """Map normalized XLSX header names to their column index."""
    mapping = {}
    for i, h in enumerate(header_cells):
        key = normalize(str(h) if h is not None else "")
        if key:
            mapping[key] = i
    return mapping


def build_line_from_row(row: tuple, specs: List[FieldSpec], header_map: dict[str, int], strict_required: bool) -> tuple[str, list[str]]:
    """Build one fixed-width line from an XLSX row and return validation issues."""
    parts: List[str] = []
    issues: list[str] = []

    for f in specs:
        key = normalize(f.name)
        idx = header_map.get(key)
        value = ""
        if idx is not None and idx < len(row):
            value = format_cell_value(row[idx])

        if strict_required and f.required.upper() == "O" and value == "":
            issues.append(f"Champ obligatoire vide: {f.name}")

        parts.append(fit_width(value, f.width))

    return "".join(parts), issues


def convert_xlsx_to_fixed(
    specs: List[FieldSpec],
    src_xlsx: Path,
    dst_txt: Path,
    source_sheet: str | None,
    report_csv: Path | None,
    strict_required: bool,
) -> None:
    """Convert an XLSX dataset to VID-HOSP fixed-width TXT using the loaded specs."""
    wb = load_workbook(src_xlsx, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Le fichier source XLSX ne contient aucun onglet.")
    ws = wb[source_sheet] if source_sheet and source_sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("Le fichier source XLSX est vide.")

    header_map = build_header_map(header)
    reports = []

    with dst_txt.open("w", encoding="utf-8", newline="\n") as out:
        for i, row in enumerate(rows, start=2):
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            line, issues = build_line_from_row(row, specs, header_map, strict_required)
            out.write(line + "\n")
            reports.append((i, " | ".join(issues) if issues else ""))

    if report_csv:
        with report_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["xlsx_row", "issues"])
            writer.writerows(reports)

    print(f"OK: {src_xlsx} -> {dst_txt} ({len(reports)} lignes exportées)")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Corrige/génère un fichier VID-HOSP à largeur fixe via formats_psy_2026_0.xlsx"
    )
    parser.add_argument(
        "--format-xlsx",
        required=True,
        help="Chemin du fichier de format (ex: formats_psy_2026_0.xlsx)",
    )
    parser.add_argument(
        "--mode",
        required=True,
        choices=["fix-txt", "xlsx-to-fixed"],
        help="fix-txt: corrige un TXT largeur fixe, xlsx-to-fixed: génère le TXT depuis un XLSX source",
    )
    parser.add_argument("--input", required=True, help="Fichier source (.txt ou .xlsx selon --mode)")
    parser.add_argument("--output", required=True, help="Fichier TXT de sortie")
    parser.add_argument(
        "--format-sheet",
        default="VID-HOSP",
        help="Nom de l'onglet format dans le fichier --format-xlsx (défaut: VID-HOSP)",
    )
    parser.add_argument(
        "--source-sheet",
        default=None,
        help="Nom de l'onglet source dans --input (mode xlsx-to-fixed). Par défaut: 1er onglet",
    )
    parser.add_argument(
        "--report",
        default=None,
        help="Chemin d'un CSV de rapport (optionnel)",
    )
    parser.add_argument(
        "--strict-required",
        action="store_true",
        help="Signale dans le rapport les champs obligatoires vides (colonne 'O').",
    )
    return parser.parse_args()


def main() -> int:
    """CLI entry point."""
    args = parse_args()

    format_xlsx = Path(args.format_xlsx).expanduser().resolve()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    report_path = Path(args.report).expanduser().resolve() if args.report else None

    if not format_xlsx.exists():
        raise SystemExit(f"Fichier format introuvable: {format_xlsx}")
    if not input_path.exists():
        raise SystemExit(f"Fichier source introuvable: {input_path}")

    specs = load_vid_hosp_spec(format_xlsx, args.format_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)

    if args.mode == "fix-txt":
        fix_txt_lines(specs, input_path, output_path, report_path)
    else:
        convert_xlsx_to_fixed(
            specs,
            input_path,
            output_path,
            args.source_sheet,
            report_path,
            args.strict_required,
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())
